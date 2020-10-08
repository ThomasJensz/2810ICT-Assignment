import openpyxl
import datetime
import matplotlib as plt


wb = openpyxl.load_workbook('Crash Statistics Victoria.xlsx')
sheet = wb['Crash Statistics Victoria']

time1 = datetime.datetime(2010, 1, 1)   #needs user input
time2 = datetime.datetime(2014, 1, 1)
crashesinhours = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]

for i in range(3,(sheet.max_row)):
    if sheet.cell(row=i, column=5).value >= time1 and sheet.cell(row=i, column=5).value <= time2:
        current = (sheet.cell(row=i,column=6).value)
        current = current.split('.')
        if current[0] == '0':
            crashesinhours[0] += 1
        if current[0] == '1':
            crashesinhours[1] += 1
        if current[0] == '2':
            crashesinhours[2] += 1
        if current[0] == '3':
            crashesinhours[3] += 1
        if current[0] == '4':
            crashesinhours[4] += 1
        if current[0] == '5':
            crashesinhours[5] += 1
        if current[0] == '6':
            crashesinhours[6] += 1
        if current[0] == '7':
            crashesinhours[7] += 1
        if current[0] == '8':
            crashesinhours[8] += 1
        if current[0] == '9':
            crashesinhours[9] += 1
        if current[0] == '10':
            crashesinhours[10] += 1
        if current[0] == '11':
            crashesinhours[11] += 1
        if current[0] == '12':
            crashesinhours[12] += 1
        if current[0] == '13':
            crashesinhours[13] += 1
        if current[0] == '14':
            crashesinhours[14] += 1
        if current[0] == '15':
            crashesinhours[15] += 1
        if current[0] == '16':
            crashesinhours[16] += 1
        if current[0] == '17':
            crashesinhours[17] += 1
        if current[0] == '18':
            crashesinhours[18] += 1
        if current[0] == '19':
            crashesinhours[19] += 1
        if current[0] == '20':
            crashesinhours[20] += 1
        if current[0] == '21':
            crashesinhours[21] += 1
        if current[0] == '22':
            crashesinhours[22] += 1
        if current[0] == '23':
            crashesinhours[23] += 1
        if current[0] == '24':
            crashesinhours[24] += 1
            
hours = list(range(1,24))
plt.plot(hours, crashesinhours)
plt.xlabel('Hour')
plt.ylabel('Number of crashes')
            
    