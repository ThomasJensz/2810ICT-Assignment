import openpyxl
import datetime
import matplotlib as plt


wb = openpyxl.load_workbook('Crash Statistics Victoria.xlsx')
sheet = wb['Crash Statistics Victoria']

time1 = datetime.datetime(2010, 1, 1)   #needs user input
time2 = datetime.datetime(2014, 1, 1)
checkDCA = input('Enter DCA code keyword: ')

for i in range(3,(sheet.max_row)):
    if sheet.cell(row=i, column=5).value >= time1 and sheet.cell(row=i, column=5).value <= time2:
        if checkDCA in (sheet.cell(row=i, column = 10).value):
            for x in range (1, 63):
                print(sheet.cell(row=i, column=x).value, end = '') 
        
        
        
        