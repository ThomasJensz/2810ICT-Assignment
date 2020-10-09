import openpyxl
import datetime
import matplotlib as plt


wb = openpyxl.load_workbook('Crash Statistics Victoria.xlsx')
sheet = wb['Crash Statistics Victoria']

for i in range(2,(sheet.max_row)):
    if sheet.cell(row=i, column=7).value == 'Yes': #alcohol involved
        if sheet.cell(row=i, column=12).value == 'Day':
            
        if sheet.cell(row=i, column=12).value == 'Dark Street lights on':
            
        if sheet.cell(row=i, column=12).value == 'Dark No street lights':
            
        if sheet.cell(row=i, column=12).value == 'Dark Street lights unknown':
            
     if sheet.cell(row=i, column=7).value == 'No': #no alcohol involved
        if sheet.cell(row=i, column=12).value == 'Day':
            
        if sheet.cell(row=i, column=12).value == 'Dark Street lights on':
            
        if sheet.cell(row=i, column=12).value == 'Dark No street lights':
            
        if sheet.cell(row=i, column=12).value == 'Dark Street lights unknown':
            
            
            
            
            #make multiple graphs to compare trends
        